import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r'D:\UNIGO\TKB toàn trường lần 4 - 12.8.xlsx', data_only=True)
ws = wb['TKB_LOP_SC']

print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# Print row 4 and 5 headers
cols_info = {}
for col in range(1, ws.max_column + 1):
    r4 = ws.cell(4, col).value
    r5 = ws.cell(5, col).value
    r3 = ws.cell(3, col).value
    cols_info[col] = (r3, r4, r5)

print("\n--- CÁC CỘT TRONG SHEET ---")
for c, (r3, r4, r5) in cols_info.items():
    if r3 or r4 or r5:
        print(f"Col {c:2d} ({openpyxl.utils.get_column_letter(c)}): r3={r3} | r4={r4} | r5={r5}")

# Search for "Nguyên" across the entire sheet
print("\n--- TẤT CẢ TIẾT CÓ 'Nguyên' ---")
found_slots = []
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r, c).value
        if v and "Nguyên" in str(v):
            # Try to find Thu, Tiet, Buoi, Lop
            found_slots.append((r, c, v))
            print(f"Row {r:2d}, Col {c:2d} ({openpyxl.utils.get_column_letter(c)}): {v}")

