import sys
import openpyxl
import docx

sys.stdout.reconfigure(encoding='utf-8')

wb_main = openpyxl.load_workbook('d:/UNIGO/TKB toàn trường CHECK - 26.8.xlsx', data_only=True)
ws_main = wb_main['TKB_LOP_SC']

col_to_info = {}
for c in range(1, ws_main.max_column + 1):
    classname = ''
    for back_c in range(c, 0, -1):
        v = ws_main.cell(4, back_c).value
        if v and str(v).strip() not in ['THỨ', 'TIẾT', '']:
            classname = str(v).strip()
            break
    sess = str(ws_main.cell(5, c).value or '').strip()
    if sess in ['Sáng', 'Chiều']:
        col_to_info[c] = (classname, sess)

days_map = ['Thứ Hai', 'Thứ Ba', 'Thứ Tư', 'Thứ Năm', 'Thứ Sáu']

# 1. Parse Nguyên
nguyen_grid_main = {}
for d_idx, start_r in enumerate([6, 11, 16, 21, 26]):
    for tiet in range(1, 6):
        r = start_r + tiet - 1
        for c, (cname, sess) in col_to_info.items():
            val = str(ws_main.cell(r, c).value or '').strip()
            if 'Nguyên' in val:
                subj = 'Tin học' if 'Tin' in val else ('Robotics' if 'Robotics' in val else 'Tin học')
                key = (d_idx, sess, tiet)
                nguyen_grid_main[key] = (subj, cname, val)

# 2. Parse Nguyệt
nguyet_grid_main = {}
for d_idx, start_r in enumerate([6, 11, 16, 21, 26]):
    for tiet in range(1, 6):
        r = start_r + tiet - 1
        for c, (cname, sess) in col_to_info.items():
            val = str(ws_main.cell(r, c).value or '').strip()
            if 'Nguyệt' in val:
                subj = 'Toán' if 'Toán' in val else ('HĐTN' if 'HĐTN' in val else val)
                key = (d_idx, sess, tiet)
                nguyet_grid_main[key] = (subj, cname, val)

print('=== NGUYÊN (Total periods:', len(nguyen_grid_main), ') ===')
for (d_idx, sess, tiet), (subj, cname, raw) in sorted(nguyen_grid_main.items()):
    print(f'{days_map[d_idx]} | {sess} Tiết {tiet} | {subj} ({cname}) [raw: {raw}]')

print('\n=== NGUYỆT (Total periods:', len(nguyet_grid_main), ') ===')
for (d_idx, sess, tiet), (subj, cname, raw) in sorted(nguyet_grid_main.items()):
    print(f'{days_map[d_idx]} | {sess} Tiết {tiet} | {subj} ({cname}) [raw: {raw}]')

# Compare Nguyen existing
wb_ng = openpyxl.load_workbook('d:/UNIGO/Thời khóa biểu giáo viên/Thời khóa biểu - Đậu Đình Nguyên.xlsx', data_only=True)
ws_ng = wb_ng.active

print('\n=== CHECK DIFFERENCE FOR NGUYỄN ===')
diff_count = 0
for sess, r_start in [('Sáng', 7), ('Chiều', 13)]:
    for tiet in range(1, 6):
        r = r_start + tiet - 1
        for d_idx in range(5):
            c = 4 + d_idx
            val_existing = str(ws_ng.cell(r, c).value or '').strip().replace('\n', ' ')
            main_item = nguyen_grid_main.get((d_idx, sess, tiet))
            val_main = f"{main_item[0]} ({main_item[1]})" if main_item else ""
            
            # normalize for comparison
            v1 = val_existing.replace('💻 ', '').replace('🤖 ', '').strip()
            v2 = val_main.strip()
            if v1 != v2:
                diff_count += 1
                print(f'DIFF: {days_map[d_idx]} | {sess} Tiết {tiet} | File: "{val_existing}" vs TKB Toàn trường: "{val_main}"')

if diff_count == 0:
    print('==> Nguyen file is 100% MATCHING TKB Toàn trường!')
else:
    print(f'==> Found {diff_count} differences!')
