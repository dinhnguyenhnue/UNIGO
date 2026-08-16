import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r'D:\UNIGO\TKB toàn trường lần 4 - 12.8.xlsx', data_only=True)
ws = wb['TKB_LOP_SC']

# Build class mapping: col -> (class_name, session)
class_map = {}
current_class = None
for c in range(1, ws.max_column + 1):
    c_name = ws.cell(4, c).value
    if c_name and str(c_name).strip() not in ('THỨ', 'TIẾT'):
        current_class = str(c_name).strip()
    session = ws.cell(5, c).value
    if session:
        session = str(session).strip()
    if current_class and session:
        class_map[c] = (current_class, session)

# Map row -> (thu, tiet)
row_map = {}
thu_names = {2: 'Thứ Hai', 3: 'Thứ Ba', 4: 'Thứ Tư', 5: 'Thứ Năm', 6: 'Thứ Sáu'}
for r in range(6, 31):
    thu_idx = (r - 6) // 5 + 2
    tiet_idx = (r - 6) % 5 + 1
    row_map[r] = (thu_names[thu_idx], tiet_idx)

# Collect all sessions for "Nguyên"
nguyen_schedule = []
for r in range(6, 31):
    thu, tiet = row_map[r]
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r, c).value
        if v and 'Nguyên' in str(v):
            cls, session = class_map.get(c, ('Unknown', 'Unknown'))
            val_str = str(v).strip()
            subject = val_str.split('-')[0].strip()
            nguyen_schedule.append({
                'thu': thu,
                'tiet': tiet,
                'session': session,
                'class': cls,
                'subject': subject,
                'raw': val_str,
                'row': r,
                'col': c
            })

print(f"Tổng số tiết của thầy Nguyên: {len(nguyen_schedule)}")
print("\n--- CHI TIẾT TỪNG TIẾT ---")
for s in nguyen_schedule:
    print(f"{s['thu']:<10} | Tiết {s['tiet']} ({s['session']:<5}) | Lớp: {s['class']:<10} | Môn: {s['subject']:<10} | [Raw: {s['raw']}]")

# Check existing files in D:\UNIGO\Thời khóa biểu giáo viên
tkb_dir = r'D:\UNIGO\Thời khóa biểu giáo viên'
print("\n--- CÁC FILE HIỆN CÓ TRONG 'Thời khóa biểu giáo viên' ---")
for f in os.listdir(tkb_dir):
    print(f" - {f}")
