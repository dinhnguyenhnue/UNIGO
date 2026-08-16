import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

# 1. Parse TKB Lần 4
wb_all = openpyxl.load_workbook(r'D:\UNIGO\TKB toàn trường lần 4 - 12.8.xlsx', data_only=True)
ws_all = wb_all['TKB_LOP_SC']

class_map = {}
current_class = None
for c in range(1, ws_all.max_column + 1):
    c_name = ws_all.cell(4, c).value
    if c_name and str(c_name).strip() not in ('THỨ', 'TIẾT'):
        current_class = str(c_name).strip()
    session = ws_all.cell(5, c).value
    if session:
        session = str(session).strip()
    if current_class and session:
        class_map[c] = (current_class, session)

thu_cols = {'Thứ Hai': 4, 'Thứ Ba': 5, 'Thứ Tư': 6, 'Thứ Năm': 7, 'Thứ Sáu': 8}
thu_names = {2: 'Thứ Hai', 3: 'Thứ Ba', 4: 'Thứ Tư', 5: 'Thứ Năm', 6: 'Thứ Sáu'}

# Matrix: (thu, session, tiet) -> (subject, class)
tkb4_matrix = {}
for r in range(6, 31):
    thu_idx = (r - 6) // 5 + 2
    tiet_idx = (r - 6) % 5 + 1
    thu = thu_names[thu_idx]
    for c in range(1, ws_all.max_column + 1):
        v = ws_all.cell(r, c).value
        if v and 'Nguyên' in str(v):
            cls, session = class_map.get(c, ('Unknown', 'Unknown'))
            val_str = str(v).strip()
            subj = 'Tin học' if 'Tin' in val_str else ('Robotics' if 'Robotics' in val_str else val_str)
            key = (thu, session, tiet_idx)
            tkb4_matrix[key] = (subj, cls)

print(f"Tổng số tiết trong TKB Lần 4: {len(tkb4_matrix)}")

# Print matrix neatly
print("\n" + "="*80)
print(f"{'BUỔI / TIẾT':<15} | {'Thứ Hai':<12} | {'Thứ Ba':<12} | {'Thứ Tư':<12} | {'Thứ Năm':<12} | {'Thứ Sáu':<12}")
print("="*80)
for session in ['Sáng', 'Chiều']:
    print(f"--- BUỔI {session.upper()} ---")
    for tiet in range(1, 6):
        row_str = f"Tiết {tiet:<10} | "
        for thu in ['Thứ Hai', 'Thứ Ba', 'Thứ Tư', 'Thứ Năm', 'Thứ Sáu']:
            cell = tkb4_matrix.get((thu, session, tiet))
            if cell:
                subj, cls = cell
                short_s = 'Tin' if subj == 'Tin học' else 'Rob'
                row_str += f"{short_s} ({cls})".ljust(12) + " | "
            else:
                row_str += " "*12 + " | "
        print(row_str)
print("="*80)
