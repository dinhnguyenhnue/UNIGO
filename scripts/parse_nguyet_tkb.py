import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('d:/UNIGO/TKB toàn trường CHECK - 26.8.xlsx', data_only=True)
ws_sc = wb['TKB_LOP_SC']

col_to_info = {}
for c in range(1, ws_sc.max_column + 1):
    classname = ''
    for back_c in range(c, 0, -1):
        v = ws_sc.cell(4, back_c).value
        if v and str(v).strip() not in ['THỨ', 'TIẾT', '']:
            classname = str(v).strip()
            break
    sess = str(ws_sc.cell(5, c).value or '').strip()
    if sess in ['Sáng', 'Chiều']:
        col_to_info[c] = (classname, sess)

days = ['Thứ Hai', 'Thứ Ba', 'Thứ Tư', 'Thứ Năm', 'Thứ Sáu']

nguyet_slots = []
for d_idx, start_r in enumerate([6, 11, 16, 21, 26]):
    day_name = days[d_idx]
    for tiet in range(1, 6):
        r = start_r + tiet - 1
        for c, (cname, sess) in col_to_info.items():
            val = str(ws_sc.cell(r, c).value or '').strip()
            if 'Nguyệt' in val:
                subj = 'Toán' if 'Toán' in val else ('HĐTN' if 'HĐTN' in val else val)
                nguyet_slots.append({
                    'day_idx': d_idx,
                    'day_name': day_name,
                    'session': sess,
                    'tiet': tiet,
                    'class': cname,
                    'subject': subj,
                    'raw': val
                })

print(f"Tổng số tiết của Cô Nguyệt: {len(nguyet_slots)} tiết/tuần\n")
print(f"{'Thứ':<10} | {'Buổi':<6} | {'Tiết':<6} | {'Lớp':<8} | {'Môn':<10} | {'Raw':<20}")
print("-" * 70)
for s in sorted(nguyet_slots, key=lambda x: (x['day_idx'], 0 if x['session']=='Sáng' else 1, x['tiet'])):
    print(f"{s['day_name']:<10} | {s['session']:<6} | Tiết {s['tiet']:<2} | {s['class']:<8} | {s['subject']:<10} | {s['raw']:<20}")

# Statistics by Subject & Class
stats_subj = {}
stats_class = {}
for s in nguyet_slots:
    stats_subj[s['subject']] = stats_subj.get(s['subject'], 0) + 1
    stats_class[s['class']] = stats_class.get(s['class'], 0) + 1

print("\n--- THỐNG KÊ THEO MÔN ---")
for subj, count in stats_subj.items():
    print(f"- {subj}: {count} tiết")

print("\n--- THỐNG KÊ THEO LỚP ---")
for cls, count in stats_class.items():
    print(f"- {cls}: {count} tiết")
