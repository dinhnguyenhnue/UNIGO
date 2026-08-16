import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r'D:\UNIGO\TKB toàn trường lần 4 - 12.8.xlsx', data_only=True)
ws = wb['TKB_LOP_SC']

print("--- ROWS (THỨ & TIẾT) ---")
for r in range(4, ws.max_row + 1):
    c1 = ws.cell(r, 1).value
    c2 = ws.cell(r, 2).value
    c11 = ws.cell(r, 11).value
    c12 = ws.cell(r, 12).value
    c21 = ws.cell(r, 21).value
    c22 = ws.cell(r, 22).value
    print(f"Row {r:2d}: Col1={c1} | Col2={c2} | Col11={c11} | Col12={c12} | Col21={c21} | Col22={c22}")
